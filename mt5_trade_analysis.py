import pandas as pd
import MetaTrader5 as mt5
from datetime import datetime, timedelta
import pytz
import os
import json
import openpyxl

def convert_to_beijing_time(timestamp):
    """将GMT时间戳转换为北京时间"""
    gmt_time = datetime.fromtimestamp(timestamp, tz=pytz.UTC)
    beijing_tz = pytz.timezone('Asia/Shanghai')
    beijing_time = gmt_time.astimezone(beijing_tz)
    return beijing_time

def get_gmt_day_range(date):
    """获取指定GMT日期的起止时间"""
    # 设置GMT时间的开始和结束
    start_time = datetime.combine(date, datetime.min.time())
    end_time = datetime.combine(date, datetime.max.time())
    
    # 将时间转换为带时区的datetime
    start_time = pytz.UTC.localize(start_time)
    end_time = pytz.UTC.localize(end_time)
    
    return start_time, end_time

def connect_mt5():
    """连接到MetaTrader5"""
    if not mt5.initialize():
        print("initialize() failed, error code =", mt5.last_error())
        return None
    return mt5

def get_trades_history(days_back=30):
    """获取交易历史数据"""
    # 定义交易数据的列名
    columns = ['ticket', 'order', 'time', 'time_msc', 'type', 'entry', 'magic', 
               'position_id', 'reason', 'volume', 'price', 'commission', 'swap', 
               'profit', 'fee', 'symbol', 'comment', 'external_id']
    
    # 计算GMT时间的日期范围
    end_date = datetime.now(pytz.UTC).date()
    start_date = end_date - timedelta(days=days_back)
    
    # 获取GMT时间起止时间点
    gmt_start, _ = get_gmt_day_range(start_date)
    _, gmt_end = get_gmt_day_range(end_date)
    
    # 获取交易历史
    deals = mt5.history_deals_get(gmt_start, gmt_end)
    
    if deals is None:
        print("No deals found")
        return None
    
    # 获取账户历史记录（入金和出金）
    account_history = mt5.history_deals_get(gmt_start, gmt_end, group="*")
    if account_history is not None:
        # 创建入金出金DataFrame
        balance_df = pd.DataFrame(list(account_history), columns=columns)
        # 只保留入金和出金记录
        balance_df = balance_df[balance_df['entry'] == mt5.DEAL_ENTRY_IN]
        balance_df = balance_df[balance_df['type'].isin([mt5.DEAL_TYPE_BALANCE, mt5.DEAL_TYPE_CREDIT])]
        balance_df['time_gmt'] = pd.to_datetime(balance_df['time'], unit='s', utc=True)
        balance_df['date'] = balance_df['time_gmt'].dt.date
    else:
        balance_df = pd.DataFrame()
    
    # 获取订单历史以获取止损止盈价格
    orders = mt5.history_orders_get(gmt_start, gmt_end)
    if orders is not None:
        # 创建订单字典，键为position_id，值为止损止盈价格
        orders_dict = {}
        for order in orders:
            if order.position_id not in orders_dict:
                orders_dict[order.position_id] = {
                    'sl': order.sl,
                    'tp': order.tp
                }
            # 如果已存在，更新为最新的止损止盈价格
            else:
                if order.sl != 0:
                    orders_dict[order.position_id]['sl'] = order.sl
                if order.tp != 0:
                    orders_dict[order.position_id]['tp'] = order.tp
    else:
        orders_dict = {}
    
    # 转换为DataFrame
    df = pd.DataFrame(list(deals), columns=columns)
    
    # 添加止损止盈价格列
    df['sl'] = df['position_id'].map(lambda x: orders_dict.get(x, {}).get('sl', None))
    df['tp'] = df['position_id'].map(lambda x: orders_dict.get(x, {}).get('tp', None))
    
    # 将时间转换为datetime并添加时区信息
    df['time_gmt'] = pd.to_datetime(df['time'], unit='s', utc=True)
    
    # 获取所有position_id的完整交易记录
    all_positions = df.groupby('position_id').agg({
        'entry': 'count',
        'profit': 'sum'
    }).reset_index()
    
    # 找出已完成的交易（有开仓和平仓记录）
    completed_positions = all_positions[all_positions['entry'] >= 2]['position_id'].tolist()
    
    # 只保留已完成交易的记录
    df = df[df['position_id'].isin(completed_positions)]
    
    # 如果没有完成的交易，返回None
    if len(df) == 0:
        print("No completed trades found")
        return None
    
    return df, balance_df

def calculate_position_times(df):
    """计算每个持仓的持仓时间和盈亏统计"""
    position_data = []
    
    for position_id in df['position_id'].unique():
        position_trades = df[df['position_id'] == position_id].sort_values('time')
        
        # 检查是否同时存在开仓和平仓记录
        open_trades = position_trades[position_trades['entry'] == mt5.DEAL_ENTRY_IN]
        close_trades = position_trades[position_trades['entry'] == mt5.DEAL_ENTRY_OUT]
        
        if len(open_trades) == 0 or len(close_trades) == 0:
            continue
        
        # 获取开仓和平仓记录
        open_trade = open_trades.iloc[0]
        close_trade = close_trades.iloc[-1]
        
        # 计算滑点
        slippage = 0.0  # 默认滑点为0
        sl_tp_price = None
        close_type = None
        
        # 检查是否有止损或止盈价格
        if pd.notna(close_trade['sl']) and close_trade['sl'] != 0:
            sl_tp_price = float(close_trade['sl'])
            close_type = 'sl'
            slippage = float(abs(close_trade['price'] - sl_tp_price))
        elif pd.notna(close_trade['tp']) and close_trade['tp'] != 0:
            sl_tp_price = float(close_trade['tp'])
            close_type = 'tp'
            slippage = float(abs(close_trade['price'] - sl_tp_price))
        else:
            # 如果没有止损止盈，则不计算滑点
            close_type = 'market'
            slippage = 0.0
        
        # 打印调试信息
        print(f"Position {position_id}:")
        print(f"Symbol: {open_trade['symbol']}")
        print(f"Type: {'Buy' if open_trade['type'] == mt5.DEAL_TYPE_BUY else 'Sell'}")
        print(f"Open price: {open_trade['price']}")
        print(f"Close price: {close_trade['price']}")
        print(f"SL: {close_trade['sl']}")
        print(f"TP: {close_trade['tp']}")
        if sl_tp_price is not None:
            print(f"SL/TP price: {sl_tp_price}")
        print(f"Slippage: {slippage}")
        print(f"Close type: {close_type}")
        print("-" * 30)
        
        position_data.append({
            'position_id': position_id,
            'open_time': position_trades['time_gmt'].min(),
            'close_time': position_trades['time_gmt'].max(),
            'symbol': open_trade['symbol'],
            'volume': float(open_trade['volume']),
            'type': open_trade['type'],
            'open_price': float(open_trade['price']),
            'close_price': float(close_trade['price']),
            'sl': float(close_trade['sl']) if pd.notna(close_trade['sl']) and close_trade['sl'] != 0 else None,
            'tp': float(close_trade['tp']) if pd.notna(close_trade['tp']) and close_trade['tp'] != 0 else None,
            'sl_tp_price': sl_tp_price,
            'slippage': slippage,
            'close_type': close_type,
            'profit': float(position_trades['profit'].sum()),
            'commission': float(position_trades['commission'].sum()),
            'swap': float(position_trades['swap'].sum())
        })
    
    # 如果没有有效的交易记录，返回空DataFrame
    if not position_data:
        return pd.DataFrame()
    
    # 创建DataFrame
    position_times = pd.DataFrame(position_data)
    
    # 计算持仓时间（秒）
    position_times['holding_time'] = (position_times['close_time'] - position_times['open_time']).dt.total_seconds()
    
    # 使用GMT时间的日期
    position_times['date'] = position_times['close_time'].dt.date
    
    # 计算总盈亏（包含手续费和隔夜费）
    position_times['total_profit'] = position_times['profit'] + position_times['commission'] + position_times['swap']
    
    # 标记盈利和亏损交易
    position_times['is_profit'] = position_times['profit'] > 0
    
    # 计算价格变动点数
    position_times['price_change'] = abs(position_times['close_price'] - position_times['open_price'])
    
    return position_times

def format_time(seconds):
    """将秒数转换为可读的时间格式"""
    return str(timedelta(seconds=int(seconds)))

def format_excel_worksheet(worksheet, df):
    """格式化Excel工作表，自动调整列宽并设置表头样式"""
    # 设置表头填充颜色（深森林绿色）
    green_fill = openpyxl.styles.PatternFill(
        start_color='005000',  # 深森林绿色
        end_color='005000',
        fill_type='solid'
    )
    
    # 设置表头字体（加粗，白色）
    header_font = openpyxl.styles.Font(
        bold=True,
        color='FFFFFF'
    )
    
    # 设置居中对齐
    center_alignment = openpyxl.styles.Alignment(
        horizontal='center',
        vertical='center'
    )
    
    # 获取数据范围
    max_row = len(df) + 1  # 加1是因为有标题行
    max_col = len(df.columns)
    
    # 应用样式到所有单元格
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.alignment = center_alignment
            
            # 如果是表头（第一行），应用表头样式
            if row == 1:
                cell.fill = green_fill
                cell.font = header_font
    
    # 自动调整列宽
    for idx, col in enumerate(df.columns):
        # 获取列名长度和该列数据的最大长度
        column_width = max(
            len(str(col)),  # 列名长度
            df[col].astype(str).map(len).max(),  # 数据最大长度
            10  # 最小宽度
        )
        # 设置列宽（稍微增加一点空间）
        worksheet.column_dimensions[openpyxl.utils.get_column_letter(idx + 1)].width = column_width + 2

def save_daily_trades(df, position_times, balance_df):
    """保存每日交易数据到单独的文件"""
    # 确保trade-data目录存在
    if not os.path.exists('trade-data'):
        os.makedirs('trade-data')
    
    # 按日期计算净值变化
    position_times_sorted = position_times.sort_values('close_time')
    initial_equity = 10000  # 初始净值设为10000
    current_equity = initial_equity
    equity_data = []
    
    # 获取所有日期（包括交易日和入金出金日）
    all_dates = pd.concat([
        pd.Series(position_times_sorted['date'].unique()),
        pd.Series(balance_df['date'].unique() if not balance_df.empty else [])
    ]).unique()
    all_dates.sort()
    
    # 按日期统计净值变化
    for date in all_dates:
        # 获取当日交易记录
        daily_trades = position_times_sorted[position_times_sorted['date'] == date]
        
        # 获取当日入金出金记录
        daily_balance = balance_df[balance_df['date'] == date] if not balance_df.empty else pd.DataFrame()
        
        # 计算当日入金出金总额
        daily_balance_change = daily_balance['profit'].sum() if not daily_balance.empty else 0
        
        # 计算当日交易盈亏
        daily_trading_profit = daily_trades['total_profit'].sum() if not daily_trades.empty else 0
        
        # 更新净值
        current_equity += daily_balance_change + daily_trading_profit
        
        # 记录当日净值数据
        equity_data.append({
            '日期': date,
            '交易盈亏': daily_trading_profit,
            '入金出金': daily_balance_change,
            '净值': current_equity,
            '权益比例': (current_equity / initial_equity - 1) * 100
        })
    
    # 创建净值DataFrame
    equity_df = pd.DataFrame(equity_data)
    
    # 按日期分组保存数据
    for date, group in position_times.groupby('date'):
        date_str = date.strftime('%Y-%m-%d')
        json_path = os.path.join('trade-data', f'{date_str}.json')
        excel_path = os.path.join('trade-data', f'{date_str}.xlsx')
        
        # 计算当日盈亏统计
        profit_trades = group[group['is_profit']]
        loss_trades = group[~group['is_profit']]
        
        # 计算盈亏比（使用平均盈利点数和平均亏损点数）
        avg_profit_points = profit_trades['price_change'].mean() if len(profit_trades) > 0 else 0
        avg_loss_points = loss_trades['price_change'].mean() if len(loss_trades) > 0 else 0
        profit_loss_ratio = avg_profit_points / avg_loss_points if avg_loss_points > 0 else 0
        
        # 准备交易数据
        trades_list = []
        for _, trade in group.iterrows():
            trade_data = {
                '交易品种': trade['symbol'],
                '交易手数': float(trade['volume']),
                '交易方向': '多单' if trade['type'] == mt5.DEAL_TYPE_BUY else '空单',
                '开仓时间(GMT)': trade['open_time'].strftime('%Y-%m-%d %H:%M:%S'),
                '平仓时间(GMT)': trade['close_time'].strftime('%Y-%m-%d %H:%M:%S'),
                '开仓价格': float(trade['open_price']),
                '平仓价格': float(trade['close_price']),
                '价格变动点数': float(trade['price_change']),
                '持仓时间': str(timedelta(seconds=int(trade['holding_time']))),
                '净利润': float(trade['profit']),
                '手续费': float(trade['commission']),
                '隔夜费': float(trade['swap']),
                '总盈亏': float(trade['total_profit'])
            }
            trades_list.append(trade_data)
        
        # 创建交易明细DataFrame
        trades_df = pd.DataFrame(trades_list)
        
        # 重命名交易明细的列
        trades_df.columns = [
            '交易品种',
            '交易手数',
            '交易方向',
            '开仓时间(GMT)',
            '平仓时间(GMT)',
            '开仓价格',
            '平仓价格',
            '价格变动点数',
            '持仓时间',
            '净利润',
            '手续费',
            '隔夜费',
            '总盈亏'
        ]
        
        # 创建汇总数据
        summary_data = {
            '总交易笔数': len(group),
            '盈利交易数': len(profit_trades),
            '亏损交易数': len(loss_trades),
            '胜率': len(profit_trades) / len(group) * 100 if len(group) > 0 else 0,
            '盈亏比': float(profit_loss_ratio),
            '平均盈利点数': float(avg_profit_points),
            '平均亏损点数': float(avg_loss_points),
            '总交易手数': float(group['volume'].sum()),
            '总盈亏': float(group['total_profit'].sum()),
            '总手续费': float(group['commission'].sum()),
            '总隔夜费': float(group['swap'].sum()),
            '多单数量': sum(group['type'] == mt5.DEAL_TYPE_BUY),
            '空单数量': sum(group['type'] == mt5.DEAL_TYPE_SELL),
            '平均持仓时间': str(timedelta(seconds=int(group['holding_time'].mean()))),
            '最长持仓时间': str(timedelta(seconds=int(group['holding_time'].max()))),
            '最短持仓时间': str(timedelta(seconds=int(group['holding_time'].min())))
        }
        
        # 创建汇总DataFrame
        summary_df = pd.DataFrame([summary_data])
        
        # 保存到JSON文件
        daily_data = {
            'date': date_str,
            'trades': trades_list,
            'summary': summary_data
        }
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(daily_data, f, ensure_ascii=False, indent=2)
        
        # 分析滑点
        slippage_trades = analyze_slippage(position_times, date)
        
        if not slippage_trades.empty:
            print(f"\n=== {date_str} 滑点订单统计 ===")
            for _, trade in slippage_trades.iterrows():
                trade_type = "多单" if trade['type'] == mt5.DEAL_TYPE_BUY else "空单"
                close_type = "止损" if trade['close_type'] == 'sl' else ("止盈" if trade['close_type'] == 'tp' else "市价")
                print(f"\n平仓时间: {trade['close_time'].strftime('%Y-%m-%d %H:%M:%S')}")
                print(f"货币对: {trade['symbol']}")
                print(f"方向: {trade_type}")
                print(f"手数: {trade['volume']:.2f}")
                print(f"平仓类型: {close_type}")
                print(f"开仓价: {trade['open_price']:.5f}")
                if trade['sl_tp_price'] is not None:
                    print(f"止损/止盈价: {trade['sl_tp_price']:.5f}")
                print(f"成交价: {trade['close_price']:.5f}")
                print(f"滑点: {trade['slippage']:.5f}")
            print("-" * 50)
        
        # 保存到Excel文件
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            # 保存交易明细
            trades_df.to_excel(writer, sheet_name='交易明细', index=False)
            format_excel_worksheet(writer.sheets['交易明细'], trades_df)
            
            # 保存每个品种的汇总数据
            symbol_summary = group.groupby('symbol').agg({
                'volume': 'sum',
                'profit': 'sum',
                'commission': 'sum',
                'swap': 'sum',
                'total_profit': 'sum',
                'price_change': ['mean', 'max', 'min'],
                'holding_time': ['mean', 'max', 'min'],
                'is_profit': ['count', 'sum']
            }).reset_index()
            
            # 计算每个品种的胜率
            symbol_summary['win_rate'] = (symbol_summary['is_profit']['sum'] / symbol_summary['is_profit']['count'] * 100).round(2)
            
            # 计算每个品种的盈亏比
            symbol_profit_loss = []
            for symbol in symbol_summary['symbol']:
                symbol_trades = group[group['symbol'] == symbol]
                profit_trades = symbol_trades[symbol_trades['is_profit']]
                loss_trades = symbol_trades[~symbol_trades['is_profit']]
                
                avg_profit = profit_trades['price_change'].mean() if len(profit_trades) > 0 else 0
                avg_loss = loss_trades['price_change'].mean() if len(loss_trades) > 0 else 0
                ratio = avg_profit / avg_loss if avg_loss > 0 else 0
                symbol_profit_loss.append(ratio)
            
            symbol_summary['profit_loss_ratio'] = symbol_profit_loss
            
            # 重命名列
            symbol_summary.columns = [
                '交易品种', 
                '总交易手数', 
                '净利润', 
                '手续费', 
                '隔夜费', 
                '总盈亏', 
                '平均点数', 
                '最大点数', 
                '最小点数',
                '平均持仓时间', 
                '最长持仓时间', 
                '最短持仓时间',
                '总交易笔数', 
                '盈利交易数', 
                '胜率', 
                '盈亏比'
            ]
            
            # 转换持仓时间格式
            symbol_summary['平均持仓时间'] = symbol_summary['平均持仓时间'].apply(lambda x: str(timedelta(seconds=int(x))))
            symbol_summary['最长持仓时间'] = symbol_summary['最长持仓时间'].apply(lambda x: str(timedelta(seconds=int(x))))
            symbol_summary['最短持仓时间'] = symbol_summary['最短持仓时间'].apply(lambda x: str(timedelta(seconds=int(x))))
            
            # 格式化数值
            for col in ['平均点数', '最大点数', '最小点数', '胜率', '盈亏比']:
                symbol_summary[col] = symbol_summary[col].round(2)
            
            symbol_summary.to_excel(writer, sheet_name='品种汇总', index=False)
            format_excel_worksheet(writer.sheets['品种汇总'], symbol_summary)
            
            # 保存日汇总数据
            summary_df.to_excel(writer, sheet_name='日汇总', index=False)
            format_excel_worksheet(writer.sheets['日汇总'], summary_df)
            
            # 保存滑点统计
            if not slippage_trades.empty:
                slippage_df = slippage_trades[[
                    'close_time', 'symbol', 'volume', 'type', 'open_price',
                    'sl_tp_price', 'close_price', 'slippage', 'close_type'
                ]].copy()
                
                # 转换时间格式
                slippage_df['close_time'] = slippage_df['close_time'].dt.strftime('%Y-%m-%d %H:%M:%S')
                
                # 转换交易方向
                slippage_df['type'] = slippage_df['type'].map({mt5.DEAL_TYPE_BUY: '多单', mt5.DEAL_TYPE_SELL: '空单'})
                
                # 转换平仓类型
                slippage_df['close_type'] = slippage_df['close_type'].map({'sl': '止损', 'tp': '止盈', 'market': '市价'})
                
                # 重命名列
                slippage_df.columns = [
                    '平仓时间(GMT)', 
                    '交易品种', 
                    '交易手数', 
                    '交易方向', 
                    '开仓价格',
                    '止损/止盈价格', 
                    '平仓价格', 
                    '滑点点数', 
                    '平仓类型'
                ]
                
                # 按时间升序排序
                slippage_df = slippage_df.sort_values('平仓时间(GMT)')
                
                # 保存到Excel
                slippage_df.to_excel(writer, sheet_name='滑点统计', index=False)
                format_excel_worksheet(writer.sheets['滑点统计'], slippage_df)
                
                # 创建滑点折线图
                workbook = writer.book
                worksheet = writer.sheets['滑点统计']
                
                # 获取数据范围
                max_row = len(slippage_df) + 1  # 加1是因为有标题行
                
                # 创建图表对象
                chart = openpyxl.chart.LineChart()
                chart.title = "滑点变化趋势"
                chart.style = 2
                
                # 设置X轴（时间）
                times = openpyxl.chart.Reference(
                    worksheet, 
                    min_col=1,  # 时间列
                    min_row=2,  # 从第二行开始（跳过标题）
                    max_row=max_row
                )
                
                # 设置Y轴（滑点）
                slippage_values = openpyxl.chart.Reference(
                    worksheet, 
                    min_col=8,  # 滑点列
                    min_row=1,  # 从第一行开始（包含标题）
                    max_row=max_row
                )
                
                # 添加数据到图表
                chart.add_data(slippage_values, titles_from_data=True)
                chart.set_categories(times)
                
                # 设置图表大小
                chart.width = 30
                chart.height = 15
                
                # 设置X轴标题
                chart.x_axis.title = "时间"
                # 设置Y轴标题
                chart.y_axis.title = "滑点点数"
                
                # 设置X轴标签的旋转角度，使其更易读
                chart.x_axis.tickLblSkip = 3  # 每隔3个标签显示一个
                chart.x_axis.tickLblPos = "low"  # 将标签位置设置在底部
                
                # 将图表添加到工作表
                worksheet.add_chart(chart, "K2")
            
            # 保存净值统计
            if not equity_df.empty:
                # 获取截至当天的净值数据
                daily_equity = equity_df[equity_df['日期'] <= date].copy()
                
                if not daily_equity.empty:
                    # 格式化数值
                    for col in ['交易盈亏', '入金出金', '净值']:
                        daily_equity[col] = daily_equity[col].round(2)
                    daily_equity['权益比例'] = daily_equity['权益比例'].round(2)
                    
                    # 保存到Excel
                    daily_equity.to_excel(writer, sheet_name='净值统计', index=False)
                    format_excel_worksheet(writer.sheets['净值统计'], daily_equity)
                    
                    # 创建净值折线图
                    workbook = writer.book
                    worksheet = writer.sheets['净值统计']
                    
                    # 获取数据范围
                    max_row = len(daily_equity) + 1  # 加1是因为有标题行
                    
                    # 创建图表对象
                    chart = openpyxl.chart.LineChart()
                    chart.title = "账户净值变化"
                    chart.style = 2
                    
                    # 设置X轴（日期）
                    dates = openpyxl.chart.Reference(
                        worksheet, 
                        min_col=1,  # 日期列
                        min_row=2,  # 从第二行开始（跳过标题）
                        max_row=max_row
                    )
                    
                    # 设置Y轴（净值）
                    values = openpyxl.chart.Reference(
                        worksheet, 
                        min_col=4,  # 净值列
                        min_row=1,  # 从第一行开始（包含标题）
                        max_row=max_row
                    )
                    
                    # 添加数据到图表
                    chart.add_data(values, titles_from_data=True)
                    chart.set_categories(dates)
                    
                    # 设置图表大小
                    chart.width = 30
                    chart.height = 15
                    
                    # 设置X轴标题
                    chart.x_axis.title = "日期"
                    # 设置Y轴标题
                    chart.y_axis.title = "净值"
                    
                    # 将图表添加到工作表
                    worksheet.add_chart(chart, "H2")
        
        print(f"已保存 {date_str} 的交易数据到 (GMT时间 00:00:00 - 23:59:59):")
        print(f"- JSON文件: {json_path}")
        print(f"- Excel文件: {excel_path}")
        print("-" * 50)

def analyze_trades_by_day(df, balance_df):
    """按天统计交易数据"""
    # 计算持仓时间
    position_times = calculate_position_times(df)
    
    # 如果没有有效的交易记录，返回空DataFrame
    if len(position_times) == 0:
        print("没有找到有效的交易记录")
        return pd.DataFrame()
    
    # 计算净值变化
    position_times_sorted = position_times.sort_values('close_time')
    initial_equity = 10000  # 初始净值设为10000
    current_equity = initial_equity
    max_equity = initial_equity
    max_drawdown = 0
    
    # 获取所有日期
    all_dates = pd.concat([
        pd.Series(position_times_sorted['date'].unique()),
        pd.Series(balance_df['date'].unique() if not balance_df.empty else [])
    ]).unique()
    all_dates.sort()
    
    # 按日期计算净值和回撤
    for date in all_dates:
        # 获取当日交易记录
        daily_trades = position_times_sorted[position_times_sorted['date'] == date]
        
        # 获取当日入金出金记录
        daily_balance = balance_df[balance_df['date'] == date] if not balance_df.empty else pd.DataFrame()
        
        # 计算当日入金出金总额
        daily_balance_change = daily_balance['profit'].sum() if not daily_balance.empty else 0
        
        # 计算当日交易盈亏
        daily_trading_profit = daily_trades['total_profit'].sum() if not daily_trades.empty else 0
        
        # 更新净值
        current_equity += daily_balance_change + daily_trading_profit
        
        # 更新最大净值和回撤
        if daily_balance_change >= 0:  # 入金不计入最大净值计算
            max_equity = max(max_equity, current_equity - daily_balance_change)
        else:  # 出金需要考虑
            max_equity = max(max_equity * (current_equity / (current_equity - daily_balance_change)), current_equity)
        
        # 计算回撤
        drawdown = (max_equity - current_equity) / max_equity * 100
        max_drawdown = max(max_drawdown, drawdown)
    
    final_equity = current_equity
    total_return = (final_equity / initial_equity - 1) * 100
    
    print("\n=== 净值统计 ===")
    print(f"初始净值: {initial_equity:.2f}")
    print(f"最终净值: {final_equity:.2f}")
    print(f"总收益率: {total_return:.2f}%")
    print(f"最大回撤: {max_drawdown:.2f}%")
    print("-" * 50)
    
    # 保存每日交易数据
    save_daily_trades(df, position_times, balance_df)
    
    # 按天和品种分组统计
    daily_stats = []
    
    for (date, symbol), group in position_times.groupby(['date', 'symbol']):
        # 计算基础统计
        stats = {
            'date': date,
            'symbol': symbol,
            'total_volume': group['volume'].sum(),
            'max_volume': group['volume'].max(),
            'min_volume': group['volume'].min(),
            'total_profit': group['total_profit'].sum(),
            'pure_profit': group['profit'].sum(),
            'commission': group['commission'].sum(),
            'swap': group['swap'].sum(),
            'avg_holding_time': group['holding_time'].mean(),
            'max_holding_time': group['holding_time'].max(),
            'min_holding_time': group['holding_time'].min(),
        }
        
        # 计算交易方向统计
        stats['type'] = {
            'buy_count': sum(group['type'] == mt5.DEAL_TYPE_BUY),
            'sell_count': sum(group['type'] == mt5.DEAL_TYPE_SELL)
        }
        
        # 计算胜率和盈亏比
        profit_trades = group[group['is_profit']]
        loss_trades = group[~group['is_profit']]
        
        stats['win_rate'] = len(profit_trades) / len(group) * 100 if len(group) > 0 else 0
        stats['avg_profit_points'] = profit_trades['price_change'].mean() if len(profit_trades) > 0 else 0
        stats['avg_loss_points'] = loss_trades['price_change'].mean() if len(loss_trades) > 0 else 0
        stats['profit_loss_ratio'] = stats['avg_profit_points'] / stats['avg_loss_points'] if stats['avg_loss_points'] > 0 else 0
        
        daily_stats.append(stats)
    
    # 创建DataFrame
    daily_summary = pd.DataFrame(daily_stats)
    
    # 将持仓时间转换为可读格式
    daily_summary['avg_holding_time'] = daily_summary['avg_holding_time'].apply(format_time)
    daily_summary['max_holding_time'] = daily_summary['max_holding_time'].apply(format_time)
    daily_summary['min_holding_time'] = daily_summary['min_holding_time'].apply(format_time)
    
    # 格式化数值
    for col in ['win_rate', 'profit_loss_ratio', 'avg_profit_points', 'avg_loss_points']:
        daily_summary[col] = daily_summary[col].round(2)
    
    return daily_summary

def format_output(daily_summary):
    """格式化输出结果"""
    print("\n=== 每日交易统计 (GMT时间 00:00:00 - 23:59:59) ===")
    for _, row in daily_summary.iterrows():
        print(f"\n日期: {row['date']}")
        print(f"品种: {row['symbol']}")
        print("\n--- 交易量统计 ---")
        print(f"总交易手数: {row['total_volume']:.2f}")
        print(f"最大单笔手数: {row['max_volume']:.2f}")
        print(f"最小单笔手数: {row['min_volume']:.2f}")
        print("\n--- 盈亏统计 ---")
        print(f"总盈亏(含费用): {row['total_profit']:.2f}")
        print(f"纯盈亏: {row['pure_profit']:.2f}")
        print(f"手续费: {row['commission']:.2f}")
        print(f"隔夜费: {row['swap']:.2f}")
        print(f"胜率: {row['win_rate']:.2f}%")
        print(f"盈亏比: {row['profit_loss_ratio']:.2f}")
        print(f"平均盈利点数: {row['avg_profit_points']:.2f}")
        print(f"平均亏损点数: {row['avg_loss_points']:.2f}")
        print("\n--- 持仓时间统计 ---")
        print(f"平均持仓时间: {row['avg_holding_time']}")
        print(f"最长持仓时间: {row['max_holding_time']}")
        print(f"最短持仓时间: {row['min_holding_time']}")
        print("\n--- 交易方向统计 ---")
        print(f"多单数量: {row['type']['buy_count']}")
        print(f"空单数量: {row['type']['sell_count']}")
        print("-" * 50)

def analyze_slippage(position_times, date):
    """分析每日滑点订单"""
    # 获取指定日期的交易记录
    daily_trades = position_times[position_times['date'] == date].copy()
    
    # 确保滑点列为数值类型
    daily_trades['slippage'] = pd.to_numeric(daily_trades['slippage'], errors='coerce')
    
    # 只保留有滑点的订单（滑点大于0）
    slippage_trades = daily_trades[daily_trades['slippage'] > 0].copy()
    
    if len(slippage_trades) == 0:
        return pd.DataFrame()
    
    # 按滑点大小排序
    slippage_trades = slippage_trades.sort_values('slippage', ascending=False)
    
    return slippage_trades  # 返回所有滑点订单，不限制数量

def main():
    # 连接MT5
    mt5_connection = connect_mt5()
    if not mt5_connection:
        return
    
    try:
        # 获取交易历史
        trades_df, balance_df = get_trades_history()
        
        if trades_df is not None:
            # 分析交易数据
            daily_analysis = analyze_trades_by_day(trades_df, balance_df)
            
            # 格式化输出结果
            format_output(daily_analysis)
            
            # 保存到CSV
            daily_analysis.to_csv('mt5_daily_trades.csv', index=False)
            print("\n统计结果已保存到 mt5_daily_trades.csv")
    
    finally:
        # 关闭MT5连接
        mt5.shutdown()

if __name__ == '__main__':
    main() 